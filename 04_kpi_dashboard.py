"""
04_kpi_dashboard.py
---------------------
Pulls together everything produced by scripts 01-03 into a single Excel
workbook (outputs/KPI_Dashboard.xlsx) with:

  - Summary sheet: the headline KPIs for your resume/interview
  - EDA sheet: per-product demand summary
  - Forecast Accuracy sheet: baseline vs ML metrics
  - Inventory Policy sheet: EOQ / safety stock / reorder point
  - Stockout Comparison sheet: before vs after

Run this LAST, after 01_eda.py, 02_forecasting.py, 03_inventory_optimization.py
have all been run at least once.
"""

import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent / "outputs"
XLSX_PATH = OUT_DIR / "KPI_Dashboard.xlsx"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_sheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 40)


def main():
    eda = pd.read_csv(OUT_DIR / "eda_summary.csv")
    model_comp = pd.read_csv(OUT_DIR / "model_comparison.csv")
    policy = pd.read_csv(OUT_DIR / "inventory_policy.csv")
    stockout = pd.read_csv(OUT_DIR / "stockout_comparison.csv")

    overall_base_mae = model_comp["baseline_MAE"].mean()
    overall_ml_mae = model_comp["ml_MAE"].mean()
    forecast_improvement = (overall_base_mae - overall_ml_mae) / overall_base_mae * 100

    total_before = stockout["stockout_days_before"].sum()
    total_after = stockout["stockout_days_after"].sum()
    stockout_reduction = (total_before - total_after) / total_before * 100 if total_before else 0

    summary = pd.DataFrame({
        "KPI": [
            "Baseline forecast MAE (avg units/day)",
            "ML forecast MAE (avg units/day)",
            "Forecast accuracy improvement (MAE basis)",
            "Baseline forecast MAPE (avg %)",
            "ML forecast MAPE (avg %)",
            "Total stockout-days -- baseline policy (60-day test, 10 SKUs)",
            "Total stockout-days -- ML-driven policy (60-day test, 10 SKUs)",
            "Stockout-day reduction",
            "Number of SKUs analyzed",
            "Historical data span",
        ],
        "Value": [
            f"{overall_base_mae:.2f}",
            f"{overall_ml_mae:.2f}",
            f"{forecast_improvement:.1f}%",
            f"{model_comp['baseline_MAPE'].mean():.2f}%",
            f"{model_comp['ml_MAPE'].mean():.2f}%",
            f"{total_before}",
            f"{total_after}",
            f"{stockout_reduction:.1f}%",
            f"{eda.shape[0]}",
            "2 years (730 days)",
        ],
    })

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        eda.to_excel(writer, sheet_name="EDA - Demand by Product", index=False)
        model_comp.to_excel(writer, sheet_name="Forecast Accuracy", index=False)
        policy.to_excel(writer, sheet_name="Inventory Policy (EOQ)", index=False)
        stockout.to_excel(writer, sheet_name="Stockout Comparison", index=False)

        for sheet_name in writer.sheets:
            style_sheet(writer.sheets[sheet_name])

    print(f"Saved KPI dashboard -> {XLSX_PATH}")
    print("\nHeadline KPIs:")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
